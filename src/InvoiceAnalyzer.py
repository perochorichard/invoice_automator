from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.ai.documentintelligence.models import AnalyzeDocumentRequest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from Invoice import Invoice

class InvoiceAnalyzer:
    def __init__(self, endpoint: str, key: str):
        self.client = DocumentIntelligenceClient(
            endpoint=endpoint, 
            credential=AzureKeyCredential(key)
        )
        self.count = 0

    def __analyze_file(self, path) -> Invoice:
        with open(path, "rb") as f:
            data = f.read()
        poller = self.client.begin_analyze_document(
            "prebuilt-invoice",
            AnalyzeDocumentRequest(bytes_source=data)
        )
        result = poller.result()
        invoice = result.documents[0]
        return Invoice(
            vendor_name=invoice.fields.get("VendorName").value_string if invoice.fields.get("VendorName") else None,
            id=invoice.fields.get("InvoiceId").value_string if invoice.fields.get("InvoiceId") else None,
            quantity=invoice.fields.get("Quantity").value_number if invoice.fields.get("Quantity") else None,
            total=invoice.fields.get("InvoiceTotal").value_currency.amount if invoice.fields.get("InvoiceTotal") and invoice.fields.get("InvoiceTotal").value_currency else None,
            invoice_date=invoice.fields.get("InvoiceDate").value_date if invoice.fields.get("InvoiceDate") else None,
            due_date=invoice.fields.get("DueDate").value_date if invoice.fields.get("DueDate") else None
        )

    def analyze(self, invoice_paths):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Invoices Compiled {self.count}"
        ws.append(["Line Item",
                   "Vendor Name",
                   "Invoice ID",
                   "Quantity",
                   "Invoice Total",
                   "Invoice Date",
                   "Due Date"])

        for idx, path in enumerate(invoice_paths):
            print(f"--Analyzing Invoice #{idx+1} {path}--")
            invoice = self.__analyze_file(path)
            ws.append([idx+1,
                       invoice.vendor_name,
                       invoice.id,
                       invoice.quantity,
                       invoice.total,
                       invoice.invoice_date,
                       invoice.due_date])
            print(f"-------------compltete---------------")

        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column  # Get the column index
            column_letter = get_column_letter(column)

            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        print("Saving to /output...")
        wb.save(f"output/invoices_compiled_{self.count}.xlsx")
        print("--PROCESS SUCCESS--")
        self.count += 1
